#!/usr/bin/env python3
"""
Starkmann Email to Excel Processor
Converts email files containing book bibliographic data to Excel format.
Supports Windows GUI and Python 3.7.2+
"""

import re
import os
import sys
import csv
import tempfile
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    HAS_TKINTER = True
except ImportError:
    HAS_TKINTER = False


class StarkmannEmailProcessor:
    """Process Starkmann email files and extract book bibliographic data."""

    COLUMNS = [
        "Authors", "Ed", "Title", "City", "Publisher", 
        "Year", "ISBN", "Format", "Currency", "Price"
    ]

    MONTH_MAP = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
        "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
        "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"
    }

    def __init__(self):
        self.data = []
        self.books_file = None

    def extract_date_from_filename(self, filename):
        """Extract date from filename in formats: YYYY_MM_DD or DD-MMM-YY."""
        basename = os.path.basename(filename)
        
        # Try YYYY_MM_DD format
        match = re.search(r'(\d{4})_(\d{2})_(\d{2})', basename)
        if match:
            return match.group(0)
        
        # Try DD-MMM-YY format
        match = re.search(r'(\d{2})-([A-Za-z]{3})-(\d{2})', basename)
        if match:
            day, month_str, year = match.groups()
            if month_str in self.MONTH_MAP:
                month_num = self.MONTH_MAP[month_str]
                full_year = f"20{year}"
                return f"{full_year}_{month_num}_{day}"
        
        return datetime.now().strftime("%Y_%m_%d")

    def process_email_file(self, input_file):
        """
        Process email file:
        1. Read with windows-1252 encoding
        2. Delete content before the first block separator line (=======...)
        3. Return cleaned content
        """
        try:
            with open(input_file, 'r', encoding='windows-1252', errors='replace') as f:
                content = f.read()
            
            # Find the start of actual book data: line with pattern ======...=(XXXX)
            pattern = r'^={64}\(\d{4}\)$'
            match = re.search(pattern, content, re.MULTILINE)
            
            if match:
                content = content[match.start():]
            
            return content
        except Exception as e:
            raise Exception(f"Error reading file: {str(e)}")

    def parse_blocks(self, content):
        """Parse content into blocks separated by ======(XXXX) lines."""
        pattern = r'^={64}\(\d{4}\)$'
        blocks = re.split(pattern, content, flags=re.MULTILINE)
        # Filter empty blocks and ensure we don't get header/footer junk
        result = []
        for b in blocks:
            cleaned = b.strip()
            if cleaned and not cleaned.startswith("Some header") and len(cleaned) > 20:
                result.append(cleaned)
        return result

    def extract_title_authors(self, lines_list):
        """
        Extract title and authors from block lines.
        Format can be multiple lines for title/editors.
        """
        if not lines_list:
            return "", "", ""
        
        # First non-empty line is title
        title = lines_list[0].strip()
        
        # Look for "ed. by" or "by" pattern for authors
        author = ""
        ed = ""
        
        # Search in first few lines
        for i, line in enumerate(lines_list[:5]):
            if "ed. by" in line.lower():
                # Extract editors
                parts = re.split(r'ed\.\s+by\s+', line, flags=re.IGNORECASE)
                if len(parts) > 1:
                    author = parts[1].strip()
                break
            elif "by" in line.lower() and "/" in line:
                parts = line.split("/")
                if len(parts) > 1:
                    author = parts[1].strip()
        
        return author, title, ed

    def extract_book_details(self, details_lines):
        """
        Extract publisher, city, year, ISBN, format, currency, price.
        Real format: City: PUBLISHER, Year.- Pages Format ISBN Price
        """
        city = ""
        publisher = ""
        year = ""
        isbn = ""
        fmt = ""
        currency = ""
        price = ""
        
        # Combine all detail lines
        details_text = ' '.join(details_lines)
        details_text = ' '.join(details_text.split())  # Normalize whitespace
        
        # Extract city and publisher (typically "City: PUBLISHER,")
        city_pub_match = re.search(r'([A-Z][a-z]+(?:\s+[a-z]+)?)\s*:\s*([A-Z][A-Z\s]+?),', details_text)
        if city_pub_match:
            city = city_pub_match.group(1).strip()
            publisher = city_pub_match.group(2).strip()
        
        # Extract year (YYYY)
        year_match = re.search(r'\b(20\d{2})\b', details_text)
        if year_match:
            year = year_match.group(1)
        
        # Extract ISBN (13-digit number)
        isbn_match = re.search(r'\b(97[89]\d{10}|\d{10}(?:\d{3})?)\b', details_text)
        if isbn_match:
            isbn = isbn_match.group(1)
        
        # Extract format (Hardback, Paperback, etc.)
        format_match = re.search(r'\b(Hardback|Paperback|Softcover|eBook|Hardcover|Cloth)\b', details_text, re.IGNORECASE)
        if format_match:
            fmt = format_match.group(1)
        
        # Extract price with currency
        price_match = re.search(r'\[(P\s*R\s*I\s*C\s*E)\]\s*([A-Z]{3})\s+([\d.]+)', details_text)
        if price_match:
            currency = price_match.group(2)
            price = price_match.group(3)
        else:
            # Try alternative format
            price_match = re.search(r'([A-Z]{3})\s+([\d.]+)(?:\s|$)', details_text)
            if price_match:
                currency = price_match.group(1)
                price = price_match.group(2)
        
        return city, publisher, year, isbn, fmt, currency, price

    def process_blocks(self, blocks):
        """Process blocks to extract book information."""
        books = []
        
        for block in blocks:
            lines = [l.strip() for l in block.split('\n') if l.strip()]
            if len(lines) < 2:
                continue
            
            # Extract title and authors from first lines
            author, title, ed = self.extract_title_authors(lines)
            
            # Extract details from subsequent lines (before empty line or price line)
            details_lines = []
            for i, line in enumerate(lines[1:], 1):
                # Stop at description or empty sections
                if line.startswith("[") or "price" in line.lower() or len(details_lines) > 5:
                    break
                details_lines.append(line)
            
            city, publisher, year, isbn, fmt, currency, price = self.extract_book_details(details_lines)
            
            # Combine multi-line title
            if len(lines) > 1 and "/" in lines[0]:
                # Multi-line title - combine first lines until we hit details
                title_lines = [lines[0]]
                for line in lines[1:]:
                    if any(c in line for c in [",", ":"]) or re.search(r'\d{4}', line):
                        break
                    if not any(x in line.lower() for x in ['ed.', 'by', 'cham', 'berlin', 'springer']):
                        title_lines.append(line)
                title = ' '.join(title_lines).strip()
            
            books.append({
                "Authors": author,
                "Ed": ed,
                "Title": title,
                "City": city,
                "Publisher": publisher,
                "Year": year,
                "ISBN": isbn,
                "Format": fmt,
                "Currency": currency,
                "Price": price
            })
        
        return books

    def save_to_csv(self, books, csv_file):
        """Save books data to CSV file."""
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=self.COLUMNS, delimiter='|')
            writer.writeheader()
            writer.writerows(books)

    def save_to_xlsx(self, books, xlsx_file):
        """Save books data to XLSX file."""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for XLSX export. Install: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Books"
        
        # Write headers
        for col_idx, col_name in enumerate(self.COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, book in enumerate(books, start=2):
            for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=book.get(col_name, ""))
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(self.COLUMNS, start=1):
            max_length = len(col_name)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
        
        wb.save(xlsx_file)

    def process(self, input_file, output_file):
        """
        Main processing pipeline:
        1. Extract date from filename
        2. Process email file
        3. Parse blocks
        4. Extract book data
        5. Save to output format (CSV or XLSX)
        """
        # Extract date for .books file naming
        date_str = self.extract_date_from_filename(input_file)
        output_dir = os.path.dirname(output_file) or "."
        self.books_file = os.path.join(output_dir, f"{date_str}.books")
        
        # Process email content
        content = self.process_email_file(input_file)
        
        # Parse into blocks
        blocks = self.parse_blocks(content)
        
        # Extract book data
        books = self.process_blocks(blocks)
        
        # Save .books file
        books_content = content
        with open(self.books_file, 'w', encoding='utf-8') as f:
            f.write(books_content)
        
        # Determine output format and save
        if output_file.lower().endswith('.xlsx'):
            self.save_to_xlsx(books, output_file)
        elif output_file.lower().endswith('.csv'):
            self.save_to_csv(books, output_file)
        else:
            # Default to XLSX
            self.save_to_xlsx(books, output_file)
        
        return len(books), self.books_file


class StarkmannGUI:
    """GUI interface for Starkmann Email Processor."""

    def __init__(self, root):
        self.root = root
        self.root.title("Starkmann Email to Excel Processor")
        self.root.geometry("600x400")
        self.root.resizable(False, False)
        
        self.processor = StarkmannEmailProcessor()
        self.input_file = tk.StringVar()
        self.output_file = tk.StringVar()
        
        self.create_widgets()

    def create_widgets(self):
        """Create GUI widgets."""
        # Title
        title_frame = ttk.Frame(self.root)
        title_frame.pack(fill=tk.X, padx=10, pady=10)
        
        title_label = ttk.Label(
            title_frame, 
            text="Starkmann Email to Excel Processor",
            font=("Helvetica", 14, "bold")
        )
        title_label.pack()
        
        # Input file section
        input_frame = ttk.LabelFrame(self.root, text="Input File", padding=10)
        input_frame.pack(fill=tk.X, padx=10, pady=5)
        
        input_entry = ttk.Entry(input_frame, textvariable=self.input_file, width=60)
        input_entry.pack(side=tk.LEFT, padx=5)
        
        input_button = ttk.Button(
            input_frame, 
            text="Browse...", 
            command=self.select_input_file
        )
        input_button.pack(side=tk.LEFT, padx=5)
        
        # Output file section
        output_frame = ttk.LabelFrame(self.root, text="Output File (*.xlsx)", padding=10)
        output_frame.pack(fill=tk.X, padx=10, pady=5)
        
        output_entry = ttk.Entry(output_frame, textvariable=self.output_file, width=60)
        output_entry.pack(side=tk.LEFT, padx=5)
        
        output_button = ttk.Button(
            output_frame, 
            text="Browse...", 
            command=self.select_output_file
        )
        output_button.pack(side=tk.LEFT, padx=5)
        
        # Process button
        process_frame = ttk.Frame(self.root)
        process_frame.pack(fill=tk.X, padx=10, pady=20)
        
        process_button = ttk.Button(
            process_frame,
            text="Process & Export",
            command=self.process_files
        )
        process_button.pack(side=tk.LEFT, padx=5, ipadx=20, ipady=10)
        
        exit_button = ttk.Button(
            process_frame,
            text="Exit",
            command=self.root.quit
        )
        exit_button.pack(side=tk.RIGHT, padx=5, ipadx=20, ipady=10)
        
        # Info section
        info_frame = ttk.LabelFrame(self.root, text="Information", padding=10)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.info_text = tk.Text(info_frame, height=10, width=70, state=tk.DISABLED)
        self.info_text.pack(fill=tk.BOTH, expand=True)

    def select_input_file(self):
        """Select input email file."""
        file = filedialog.askopenfilename(
            title="Select Email File",
            filetypes=[("Text files", "*.txt"), ("Email files", "*.eml"), ("All files", "*.*")]
        )
        if file:
            self.input_file.set(file)
            self.auto_set_output_filename(file)

    def select_output_file(self):
        """Select output Excel file."""
        file = filedialog.asksaveasfilename(
            title="Save Excel File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file:
            self.output_file.set(file)

    def auto_set_output_filename(self, input_file):
        """Auto-generate output filename based on input filename."""
        basename = os.path.basename(input_file)
        date_match = re.search(r'(\d{4}_\d{2}_\d{2}|\d{2}-[A-Za-z]{3}-\d{2})', basename)
        
        if date_match:
            date_str = self.processor.extract_date_from_filename(input_file)
        else:
            date_str = datetime.now().strftime("%Y_%m_%d")
        
        output_dir = os.path.dirname(input_file)
        output_filename = os.path.join(output_dir, f"{date_str}.xlsx")
        self.output_file.set(output_filename)

    def append_info(self, text):
        """Append text to info display."""
        self.info_text.config(state=tk.NORMAL)
        self.info_text.insert(tk.END, text + "\n")
        self.info_text.see(tk.END)
        self.info_text.config(state=tk.DISABLED)
        self.root.update()

    def clear_info(self):
        """Clear info display."""
        self.info_text.config(state=tk.NORMAL)
        self.info_text.delete(1.0, tk.END)
        self.info_text.config(state=tk.DISABLED)

    def process_files(self):
        """Process input file and export to Excel."""
        self.clear_info()
        
        # Validation
        if not self.input_file.get():
            messagebox.showerror("Error", "Please select an input file.")
            return
        
        if not self.output_file.get():
            messagebox.showerror("Error", "Please specify an output file.")
            return
        
        if not os.path.exists(self.input_file.get()):
            messagebox.showerror("Error", f"Input file not found: {self.input_file.get()}")
            return
        
        try:
            self.append_info("Starting processing...")
            self.append_info(f"Input: {self.input_file.get()}")
            self.append_info(f"Output: {self.output_file.get()}")
            self.append_info("---")
            
            # Process
            num_books, books_file = self.processor.process(
                self.input_file.get(),
                self.output_file.get()
            )
            
            self.append_info(f"✓ Processing completed successfully!")
            self.append_info(f"✓ Books extracted: {num_books}")
            self.append_info(f"✓ Excel file saved: {self.output_file.get()}")
            self.append_info(f"✓ Books file saved: {books_file}")
            self.append_info("---")
            self.append_info("Done!")
            
            messagebox.showinfo(
                "Success",
                f"Processing completed!\n\n"
                f"Books extracted: {num_books}\n"
                f"Excel file: {self.output_file.get()}\n"
                f"Books file: {books_file}"
            )
        
        except Exception as e:
            error_msg = f"Error: {str(e)}"
            self.append_info(error_msg)
            messagebox.showerror("Processing Error", error_msg)


def main():
    """Main entry point."""
    # Check dependencies
    if not HAS_TKINTER:
        print("Error: tkinter is required but not installed.")
        print("On Windows, install Python with tkinter included.")
        sys.exit(1)
    
    if not HAS_OPENPYXL:
        print("Warning: openpyxl not installed. XLSX export may not work.")
        print("Install: pip install openpyxl")
    
    root = tk.Tk()
    gui = StarkmannGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
