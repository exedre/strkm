#!/usr/bin/env python3
"""
Test suite for Starkmann Email Processor
"""

import os
import sys
import tempfile
from pathlib import Path

# Import the processor
sys.path.insert(0, os.path.dirname(__file__))
from starkmann_email_processor import StarkmannEmailProcessor


def test_date_extraction():
    """Test date extraction from filenames."""
    processor = StarkmannEmailProcessor()
    
    # Test YYYY_MM_DD format
    date = processor.extract_date_from_filename("file_2024_11_25.txt")
    assert date == "2024_11_25", f"Expected 2024_11_25, got {date}"
    print("✓ YYYY_MM_DD format test passed")
    
    # Test DD-MMM-YY format
    date = processor.extract_date_from_filename("25-Nov-24.txt")
    assert date == "2024_11_25", f"Expected 2024_11_25, got {date}"
    print("✓ DD-MMM-YY format test passed")


def test_title_authors_extraction():
    """Test extraction of title and authors."""
    processor = StarkmannEmailProcessor()
    
    # Test with all 3 parts
    author, title, ed = processor.extract_title_authors("01|John Smith; Jane Doe|Introduction to Python")
    assert author == "John Smith; Jane Doe", f"Author extraction failed: {author}"
    assert title == "Introduction to Python", f"Title extraction failed: {title}"
    assert ed == "01", f"Edition extraction failed: {ed}"
    print("✓ Title/Authors extraction test passed")


def test_book_details_extraction():
    """Test extraction of book details."""
    processor = StarkmannEmailProcessor()
    
    details_text = "Berlin|Springer|2024|978-3-031-12345-6|Hardback|EUR|49.99"
    city, publisher, year, isbn, fmt, currency, price = processor.extract_book_details(details_text)
    
    assert city == "Berlin", f"City extraction failed: {city}"
    assert publisher == "Springer", f"Publisher extraction failed: {publisher}"
    assert year == "2024", f"Year extraction failed: {year}"
    assert isbn == "978-3-031-12345-6", f"ISBN extraction failed: {isbn}"
    print("✓ Book details extraction test passed")


def test_sample_processing():
    """Test processing with sample data."""
    processor = StarkmannEmailProcessor()
    
    # Create sample email content
    sample_content = """================================================================(0001)
Author One|Great Book Title|1st Ed
Berlin|Springer|2024|978-3-031-12345-6|Hardback|EUR|49.99

================================================================(0002)
Author Two|Another Great Book|2nd Ed
New York|Penguin|2023|978-0-14-118-345-2|Paperback|USD|29.99"""
    
    # Parse blocks using the processor's method
    blocks = processor.parse_blocks(sample_content)
    
    assert len(blocks) == 2, f"Expected 2 blocks, got {len(blocks)}: {blocks}"
    print(f"✓ Block parsing test passed ({len(blocks)} blocks found)")
    
    # Process blocks
    books = processor.process_blocks(blocks)
    assert len(books) == 2, f"Expected 2 books, got {len(books)}"
    
    # Verify first book
    book1 = books[0]
    assert book1["Authors"] == "Author One", f"First book author wrong: {book1['Authors']}"
    assert book1["Title"] == "Great Book Title", f"First book title wrong: {book1['Title']}"
    
    print("✓ Sample processing test passed")
    print(f"  Book 1: {book1['Title']} by {book1['Authors']}")
    print(f"  Book 2: {books[1]['Title']} by {books[1]['Authors']}")


def test_csv_export():
    """Test CSV export."""
    processor = StarkmannEmailProcessor()
    
    # Create test books
    books = [
        {
            "Authors": "Test Author",
            "Ed": "1st",
            "Title": "Test Book",
            "City": "Berlin",
            "Publisher": "Test Pub",
            "Year": "2024",
            "ISBN": "123-456",
            "Format": "Hardback",
            "Currency": "EUR",
            "Price": "29.99"
        }
    ]
    
    # Export to temporary CSV
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        csv_file = f.name
    
    try:
        processor.save_to_csv(books, csv_file)
        
        # Verify file was created
        assert os.path.exists(csv_file), "CSV file not created"
        
        # Check content
        with open(csv_file, 'r') as f:
            content = f.read()
            assert "Test Author" in content, "Author not found in CSV"
            assert "Test Book" in content, "Title not found in CSV"
        
        print("✓ CSV export test passed")
    finally:
        if os.path.exists(csv_file):
            os.unlink(csv_file)


def test_xlsx_export():
    """Test XLSX export (if openpyxl available)."""
    try:
        import openpyxl
    except ImportError:
        print("⚠ openpyxl not installed, skipping XLSX test")
        return
    
    processor = StarkmannEmailProcessor()
    
    # Create test books
    books = [
        {
            "Authors": "Test Author",
            "Ed": "1st",
            "Title": "Test Book",
            "City": "Berlin",
            "Publisher": "Test Pub",
            "Year": "2024",
            "ISBN": "123-456",
            "Format": "Hardback",
            "Currency": "EUR",
            "Price": "29.99"
        }
    ]
    
    # Export to temporary XLSX
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        xlsx_file = f.name
    
    try:
        processor.save_to_xlsx(books, xlsx_file)
        
        # Verify file was created
        assert os.path.exists(xlsx_file), "XLSX file not created"
        
        # Check content using openpyxl
        wb = openpyxl.load_workbook(xlsx_file)
        ws = wb.active
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "Authors" in headers, "Authors header not found"
        assert "Title" in headers, "Title header not found"
        
        # Check data
        assert ws['A2'].value == "Test Author", "Author data not found"
        
        print("✓ XLSX export test passed")
    finally:
        if os.path.exists(xlsx_file):
            os.unlink(xlsx_file)


def main():
    """Run all tests."""
    print("=" * 60)
    print("Starkmann Email Processor - Test Suite")
    print("=" * 60)
    print()
    
    try:
        test_date_extraction()
        test_title_authors_extraction()
        test_book_details_extraction()
        test_sample_processing()
        test_csv_export()
        test_xlsx_export()
        
        print()
        print("=" * 60)
        print("✓ All tests passed!")
        print("=" * 60)
        return 0
    
    except AssertionError as e:
        print()
        print("=" * 60)
        print(f"✗ Test failed: {str(e)}")
        print("=" * 60)
        return 1
    
    except Exception as e:
        print()
        print("=" * 60)
        print(f"✗ Unexpected error: {str(e)}")
        print("=" * 60)
        return 1


if __name__ == "__main__":
    sys.exit(main())
